"""
BRVM Data Extractor v3
Extraction automatique des données actions — Bulletins Officiels de la Cote BRVM
Supporte l'extraction sur années antérieures complètes avec détection intelligente des séances.
v3: Ajout du mode Intervalle de dates (date début → date fin) avec export Excel intégré.
"""

import streamlit as st
import pandas as pd
import requests
import pdfplumber
import re
import io
import time
import threading
import queue
from datetime import date, datetime, timedelta
from pathlib import Path
import warnings
warnings.filterwarnings('ignore')

# ══════════════════════════════════════════════════════════════════════════════
# CONFIG & CONSTANTES
# ══════════════════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="BRVM Data Extractor v3",
    page_icon="📈",
    layout="wide",
    initial_sidebar_state="expanded"
)

BASE_URL       = "https://bfin.brvm.org/boc/"
BOC_LIST_URL   = BASE_URL + "boc_jour.aspx"
PDF_TEMPLATE   = BASE_URL + "BOC_JOUR/BOC_{code}.pdf"
DATA_DIR       = Path("brvm_data")
DATA_DIR.mkdir(exist_ok=True)
CACHE_FILE     = DATA_DIR / "extracted_data.parquet"
LOG_FILE       = DATA_DIR / "extraction_log.json"

SECTORS = {'TEL','FIN','CD','CB','IND','ENE','SPU'}
SECTOR_NAMES = {
    'TEL': 'Télécommunications',
    'FIN': 'Services Financiers',
    'CD':  'Consommation Discrétionnaire',
    'CB':  'Consommation de Base',
    'IND': 'Industriels',
    'ENE': 'Énergie',
    'SPU': 'Services Publics',
}

ALL_SYMBOLS = [
    "ABJC","BICB","BICC","BNBC","BOAB","BOABF","BOAC","BOAM","BOAN","BOAS",
    "CABC","CBIBF","CFAC","CIEC","ECOC","ETIT","FTSC","LNBB","NEIC","NSBC",
    "NTLC","ONTBF","ORAC","ORGT","PALC","PRSC","SAFC","SCRC","SDCC","SDSC",
    "SGBC","SHEC","SIBC","SICC","SIVC","SLBC","SMBC","SNTS","SOGC","SPHC",
    "STAC","STBC","TTLC","TTLS","UNLC","UNXC",
]

# Jours fériés fixes UEMOA (mois, jour)
UEMOA_FIXED_HOLIDAYS = {
    (1,1),(1,2),(5,1),(8,15),(11,1),(12,24),(12,25),(12,26),
}

# Années disponibles (BRVM numérique depuis ~2015)
FIRST_YEAR = 2015
CURRENT_YEAR = datetime.now().year

# ══════════════════════════════════════════════════════════════════════════════
# STYLES
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Space+Mono:ital,wght@0,400;0,700;1,400&family=DM+Sans:wght@300;400;500;600;700&display=swap');

:root {
    --gold:        #C9A84C;
    --gold-light:  #E8C96B;
    --gold-dim:    rgba(201,168,76,0.15);
    --dark:        #080C18;
    --surface:     #0E1525;
    --surface2:    #172035;
    --surface3:    #1E2A40;
    --border:      rgba(201,168,76,0.18);
    --text:        #DDE3EF;
    --muted:       #7A8699;
    --green:       #2ECC71;
    --red:         #E74C3C;
    --blue:        #3B82F6;
    --orange:      #F39C12;
}

html, body, .stApp { background: var(--dark) !important; }

/* ── Sidebar ── */
section[data-testid="stSidebar"] {
    background: var(--surface) !important;
    border-right: 1px solid var(--border) !important;
}
section[data-testid="stSidebar"] .block-container { padding: 1rem 1rem 2rem; }

/* ── Typography ── */
h1,h2,h3,h4,h5 { font-family:'Space Mono',monospace !important; }
h1  { color: var(--gold)!important; font-size:1.7rem!important; letter-spacing:-.02em; }
h2  { color: var(--text)!important; font-size:1.15rem!important; }
h3  { color: var(--gold-light)!important; font-size:.95rem!important; }
p, div, span, label, li { font-family:'DM Sans',sans-serif!important; color:var(--text); }

/* ── Metrics ── */
[data-testid="metric-container"] {
    background: var(--surface2)!important;
    border: 1px solid var(--border)!important;
    border-radius: 10px!important;
    padding: 14px 18px!important;
}
[data-testid="stMetricLabel"]  { color:var(--muted)!important; font-size:.72rem!important; text-transform:uppercase; letter-spacing:.09em; }
[data-testid="stMetricValue"]  { color:var(--gold)!important; font-family:'Space Mono',monospace!important; font-size:1.35rem!important; }
[data-testid="stMetricDelta"]  { font-size:.82rem!important; }

/* ── Buttons ── */
.stButton>button {
    background: linear-gradient(135deg,#C9A84C,#9A6F1E)!important;
    color: #080C18!important;
    font-family:'Space Mono',monospace!important;
    font-weight:700!important;
    font-size:.78rem!important;
    border:none!important;
    border-radius:7px!important;
    padding:10px 22px!important;
    letter-spacing:.06em!important;
    text-transform:uppercase!important;
    transition:all .2s!important;
}
.stButton>button:hover { transform:translateY(-2px); box-shadow:0 6px 24px rgba(201,168,76,.35)!important; }
.stButton>button:disabled { opacity:.4!important; }

/* ── Danger button ── */
.btn-danger>button {
    background: linear-gradient(135deg,#E74C3C,#A93226)!important;
    color: #fff!important;
}

/* ── Tabs ── */
.stTabs [data-baseweb="tab-list"] {
    background: var(--surface)!important;
    border-bottom: 1px solid var(--border)!important;
    gap: 4px;
}
.stTabs [data-baseweb="tab"] {
    color: var(--muted)!important;
    font-family:'Space Mono',monospace!important;
    font-size:.74rem!important;
    padding: 10px 16px!important;
    border-radius: 6px 6px 0 0!important;
}
.stTabs [aria-selected="true"] {
    color: var(--gold)!important;
    background: var(--surface2)!important;
    border-bottom: 2px solid var(--gold)!important;
}
.stTabs [data-baseweb="tab-panel"] { background:var(--dark)!important; padding-top:20px!important; }

/* ── DataFrames ── */
[data-testid="stDataFrameResizable"] { background:var(--surface2)!important; border-radius:8px!important; }
.stDataFrame { border: 1px solid var(--border)!important; border-radius:8px!important; }

/* ── Inputs ── */
.stSelectbox>div>div, .stMultiSelect>div>div {
    background:var(--surface2)!important; border:1px solid var(--border)!important; border-radius:7px!important;
}
.stTextInput>div>div>input {
    background:var(--surface2)!important; border:1px solid var(--border)!important;
    color:var(--text)!important; border-radius:7px!important;
}
.stSlider>div>div>div { color:var(--text)!important; }
[data-testid="stWidgetLabel"] { color:var(--muted)!important; font-size:.78rem!important; }

/* ── Progress ── */
.stProgress>div>div>div { background:var(--gold)!important; }

/* ── Expander ── */
.streamlit-expanderHeader {
    background:var(--surface2)!important; border:1px solid var(--border)!important;
    font-family:'Space Mono',monospace!important; color:var(--gold)!important; border-radius:7px!important;
}
.streamlit-expanderContent { background:var(--surface3)!important; border:1px solid var(--border)!important; }

/* ── Alerts ── */
.stSuccess,.stInfo,.stWarning,.stError { border-radius:8px!important; }

/* ── Custom components ── */
.brvm-card {
    background: var(--surface2);
    border: 1px solid var(--border);
    border-radius: 10px;
    padding: 20px;
    margin-bottom: 14px;
}
.brvm-header {
    display:flex; align-items:center; gap:14px;
    margin-bottom:22px; padding-bottom:16px;
    border-bottom: 1px solid var(--border);
}
.badge {
    display:inline-block; padding:3px 9px; border-radius:5px;
    font-size:.68rem; font-weight:700; letter-spacing:.07em;
    font-family:'Space Mono',monospace;
}
.badge-up   { background:rgba(46,204,113,.15); color:#2ECC71; border:1px solid rgba(46,204,113,.3); }
.badge-down { background:rgba(231,76,60,.15);  color:#E74C3C; border:1px solid rgba(231,76,60,.3); }
.badge-flat { background:rgba(201,168,76,.12); color:var(--gold); border:1px solid var(--border); }
.badge-info { background:rgba(59,130,246,.15); color:#60A5FA; border:1px solid rgba(59,130,246,.3); }

.dot { width:8px;height:8px;border-radius:50%;display:inline-block;margin-right:5px; }
.dot-g { background:#2ECC71;box-shadow:0 0 6px #2ECC71; }
.dot-r { background:#E74C3C;box-shadow:0 0 6px #E74C3C; }
.dot-o { background:#F39C12; }
.dot-b { background:#3B82F6;box-shadow:0 0 6px #3B82F6; }

.ticker-row {
    display:flex;justify-content:space-between;align-items:center;
    padding:7px 12px; border-radius:7px; margin-bottom:3px;
}
.ticker-row:hover { background:rgba(201,168,76,.06); }

.year-card {
    background: var(--surface3);
    border: 1px solid var(--border);
    border-radius: 8px;
    padding: 12px 16px;
    margin-bottom: 8px;
    cursor: pointer;
    transition: all .15s;
}
.year-card:hover { border-color: var(--gold); background: var(--surface2); }
.year-card.active { border-color: var(--gold); background: rgba(201,168,76,.08); }

.progress-block {
    background: var(--surface3);
    border: 1px solid var(--border);
    border-radius: 8px;
    padding: 16px;
    margin: 8px 0;
    font-family: 'Space Mono', monospace;
    font-size: .78rem;
    color: var(--muted);
    max-height: 280px;
    overflow-y: auto;
}
.log-ok   { color: #2ECC71; }
.log-warn { color: #F39C12; }
.log-err  { color: #E74C3C; }
.log-info { color: #60A5FA; }

hr.divider { border:none; border-top:1px solid var(--border); margin:18px 0; }
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# FONCTIONS UTILITAIRES
# ══════════════════════════════════════════════════════════════════════════════

def clean_num(s):
    if s is None: return None
    s = str(s).strip().replace('\n',' ').replace('\xa0','').replace(' ','')
    s = s.replace(',','.').replace('%','').replace('+','')
    s = re.sub(r'[^\d.\-]','',s)
    try: return float(s) if s and s not in ('.','−','-') else None
    except: return None

def clean_pct(s):
    if s is None: return None
    s = str(s).strip().replace('\n',' ').replace(' ','').replace(',','.').replace('%','')
    try: return float(s)
    except: return None

def fmt_num(n, suffix=''):
    if n is None or (isinstance(n,float) and pd.isna(n)): return '—'
    if abs(n)>=1e12: return f"{n/1e12:.2f} Tn{suffix}"
    if abs(n)>=1e9:  return f"{n/1e9:.2f} Md{suffix}"
    if abs(n)>=1e6:  return f"{n/1e6:.2f} M{suffix}"
    if abs(n)>=1e3:  return f"{n/1e3:.1f} K{suffix}"
    return f"{n:,.0f}{suffix}"

def badge(val, suffix='%'):
    if val is None or (isinstance(val,float) and pd.isna(val)):
        return '<span class="badge badge-flat">—</span>'
    if val>0:  return f'<span class="badge badge-up">▲ {val:+.2f}{suffix}</span>'
    if val<0:  return f'<span class="badge badge-down">▼ {val:.2f}{suffix}</span>'
    return f'<span class="badge badge-flat">= 0.00{suffix}</span>'

def date_code(d: date) -> str:
    return d.strftime('%Y%m%d')

def code_to_display(c: str) -> str:
    return f"{c[6:8]}/{c[4:6]}/{c[:4]}"

def display_to_code(s: str) -> str:
    return s[6:10]+s[3:5]+s[0:2]


# ══════════════════════════════════════════════════════════════════════════════
# GÉNÉRATION DES DATES CALENDAIRES
# ══════════════════════════════════════════════════════════════════════════════

def get_candidate_dates(year: int) -> list[str]:
    """
    Retourne tous les codes YYYYMMDD des jours Lundi-Vendredi de l'année,
    en excluant les jours fériés fixes UEMOA.
    """
    codes = []
    d = date(year, 1, 1)
    end = date(year, 12, 31)
    while d <= end:
        if d.weekday() < 5 and (d.month, d.day) not in UEMOA_FIXED_HOLIDAYS:
            codes.append(date_code(d))
        d += timedelta(days=1)
    return codes


def get_candidate_dates_range(year: int, month_from: int, month_to: int) -> list[str]:
    """Retourne les codes pour une plage mois dans une année."""
    codes = []
    start = date(year, month_from, 1)
    # Last day of month_to
    if month_to == 12:
        end = date(year, 12, 31)
    else:
        end = date(year, month_to+1, 1) - timedelta(days=1)
    d = start
    while d <= end:
        if d.weekday() < 5 and (d.month, d.day) not in UEMOA_FIXED_HOLIDAYS:
            codes.append(date_code(d))
        d += timedelta(days=1)
    return codes


def get_candidate_dates_interval(date_start: date, date_end: date) -> list[str]:
    """
    Retourne tous les codes YYYYMMDD des jours Lun–Ven entre date_start et date_end (inclus),
    en excluant les jours fériés fixes UEMOA.
    Couvre plusieurs années si besoin.
    """
    if date_start > date_end:
        return []
    codes = []
    d = date_start
    while d <= date_end:
        if d.weekday() < 5 and (d.month, d.day) not in UEMOA_FIXED_HOLIDAYS:
            codes.append(date_code(d))
        d += timedelta(days=1)
    return codes


# ══════════════════════════════════════════════════════════════════════════════
# SCRAPING & PARSING
# ══════════════════════════════════════════════════════════════════════════════

@st.cache_data(ttl=1800, show_spinner=False)
def fetch_recent_boc_list() -> list[str]:
    """Scrape la page BRVM pour obtenir les codes des BOC récents."""
    try:
        r = requests.get(BOC_LIST_URL, timeout=15,
                         headers={'User-Agent':'Mozilla/5.0'})
        r.raise_for_status()
        codes = re.findall(r'BOC_(\d{8})\.pdf', r.text)
        return list(dict.fromkeys(codes))  # deduplicate preserving order
    except:
        return []


def download_pdf(code: str, timeout: int = 25) -> bytes | None:
    """Télécharge un PDF BOC. Retourne les bytes ou None si indisponible."""
    url = PDF_TEMPLATE.format(code=code)
    try:
        r = requests.get(url, timeout=timeout,
                         headers={'User-Agent':'Mozilla/5.0'})
        if r.status_code == 200 and r.headers.get('Content-Type','').startswith('application/pdf'):
            if len(r.content) > 30_000:   # PDF réel > 30 KB
                return r.content
        return None
    except:
        return None


def parse_actions(pdf_bytes: bytes, date_str: str) -> list[dict]:
    """
    Extrait toutes les lignes actions d'un PDF BOC.
    Gère les formats à 14, 15, 16 et 17 colonnes selon les années.
    """
    rows = []
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages:
                for table in (page.extract_tables() or []):
                    if not table or len(table) < 2:
                        continue
                    for row in table:
                        if not row or len(row) < 8:
                            continue
                        sect = str(row[0]).strip().upper() if row[0] else ''
                        sym  = str(row[1]).strip().upper() if row[1] else ''
                        if sect not in SECTORS:
                            continue
                        if not sym or not re.match(r'^[A-Z]{3,8}$', sym):
                            continue
                        titre = str(row[2]).strip().replace('\n',' ') if row[2] else ''
                        if not titre:
                            continue

                        n = len(row)
                        # Format classique 16 col (2023+):
                        # [sect, sym, titre, '', ouv, ouv2, clo, var_j, vol, val, ref, var_an, div, div_date, rdt, per]
                        # Format ancien 14 col:
                        # [sect, sym, titre, ouv, clo, var_j, vol, val, ref, var_an, div, div_date, rdt, per]

                        if n >= 16:
                            ouv    = clean_num(row[4])
                            clo    = clean_num(row[6])
                            var_j  = clean_pct(row[7])
                            vol    = clean_num(row[8])
                            val    = clean_num(row[9])
                            ref    = clean_num(row[10])
                            var_an = clean_pct(row[11])
                            div    = clean_num(row[12])
                            div_d  = str(row[13]).strip() if row[13] else ''
                            rdt    = clean_pct(row[14])
                            per    = clean_num(row[15])
                        elif n >= 14:
                            ouv    = clean_num(row[3])
                            clo    = clean_num(row[4])
                            var_j  = clean_pct(row[5])
                            vol    = clean_num(row[6])
                            val    = clean_num(row[7])
                            ref    = clean_num(row[8])
                            var_an = clean_pct(row[9])
                            div    = clean_num(row[10])
                            div_d  = str(row[11]).strip() if row[11] else ''
                            rdt    = clean_pct(row[12])
                            per    = clean_num(row[13])
                        else:
                            continue

                        if ouv is None and clo is None:
                            continue

                        rows.append({
                            'date':             date_str,
                            'secteur_code':     sect,
                            'secteur':          SECTOR_NAMES.get(sect, sect),
                            'symbole':          sym,
                            'titre':            titre,
                            'ouverture':        ouv,
                            'cloture':          clo,
                            'variation_jour_pct': var_j,
                            'volume':           vol,
                            'valeur_fcfa':      val,
                            'cours_reference':  ref,
                            'variation_annee_pct': var_an,
                            'dividende_net':    div,
                            'date_dividende':   div_d,
                            'rendement_net_pct': rdt,
                            'per':              per,
                        })
    except Exception:
        pass
    return rows


# ══════════════════════════════════════════════════════════════════════════════
# CACHE PERSISTANT
# ══════════════════════════════════════════════════════════════════════════════

def load_cache() -> pd.DataFrame | None:
    if CACHE_FILE.exists():
        try:
            return pd.read_parquet(CACHE_FILE)
        except:
            return None
    return None


def save_cache(df: pd.DataFrame):
    df.to_parquet(CACHE_FILE, index=False)


def merge_into_cache(existing: pd.DataFrame | None, new_rows: list[dict]) -> pd.DataFrame:
    if not new_rows:
        return existing if existing is not None else pd.DataFrame()
    new_df = pd.DataFrame(new_rows)
    if existing is None or existing.empty:
        return new_df
    combined = pd.concat([existing, new_df], ignore_index=True)
    combined = combined.drop_duplicates(subset=['date','symbole'], keep='last')
    return combined.sort_values(['date','symbole']).reset_index(drop=True)


def load_log() -> dict:
    if LOG_FILE.exists():
        import json
        try:
            return __import__('json').loads(LOG_FILE.read_text())
        except:
            return {}
    return {}


def save_log(log: dict):
    import json
    LOG_FILE.write_text(json.dumps(log, indent=2))


def generate_excel_workbook(data_df: pd.DataFrame, label: str = "") -> bytes:
    """
    Génère un workbook Excel complet à partir d'un DataFrame BRVM.

    Structure des feuilles :
      1. Tableau Complet  — format wide : 1 ligne/date, colonnes groupées par symbole
                            (en-têtes fusionnées sur 2 niveaux : symbole → champ)
                            Colonnes : Date | ABJC [Ouv, Clo, Var%, Vol, Val, VarAn%, RdtNet%, PER] | BICB …
      2. Pivot Clôture    — 1 ligne/date, 1 col/symbole, valeur = cours clôture
      3. Pivot Volumes    — 1 ligne/date, 1 col/symbole, valeur = volume échangé
      4. Données Brutes   — format long classique, 1 ligne par (date, symbole)
      5. Stats par Titre  — synthèse par symbole sur la période
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── Palette ────────────────────────────────────────────────────────────────
    C_DARK      = '0E1525'
    C_GOLD      = 'C9A84C'
    C_GOLD_DARK = '9A6F1E'
    C_WHITE     = 'FFFFFF'
    C_LIGHT_BG  = 'F4F6FA'
    C_EVEN      = 'EEF1F8'
    C_SYM_HDR   = ['1A2A4A', '17263F', '142235', '112030', '0F1D2C']  # alternance sombres

    def _font(bold=False, color=C_WHITE, size=9, name='Calibri'):
        return Font(bold=bold, color=color, size=size, name=name)

    def _fill(color):
        return PatternFill('solid', start_color=color, end_color=color)

    def _border(color='CCCCCC'):
        s = Side(style='thin', color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    def _center():
        return Alignment(horizontal='center', vertical='center', wrap_text=False)

    thin_brd  = _border()
    thick_brd = _border('888888')

    # ── Champs détaillés à afficher pour chaque symbole (ordre fixé) ──────────
    FIELDS = [
        ('ouverture',           'Ouv.',     '#,##0',   8),
        ('cloture',             'Clô.',     '#,##0',   8),
        ('variation_jour_pct',  'Var.J%',   '+0.00%;-0.00%;0.00%', 7),
        ('volume',              'Volume',   '#,##0',   9),
        ('valeur_fcfa',         'Val.FCFA', '#,##0',   11),
        ('variation_annee_pct', 'VarAn%',   '+0.00%;-0.00%;0.00%', 7),
        ('rendement_net_pct',   'Rdt%',     '0.00%',   6),
        ('per',                 'PER',      '0.00',    6),
    ]
    N_FIELDS = len(FIELDS)

    wb = openpyxl.Workbook()

    # ══════════════════════════════════════════════════════════════════════════
    # FEUILLE 1 — TABLEAU COMPLET (format wide, en-têtes fusionnées sur 2 lignes)
    # ══════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Tableau Complet"
    ws1.sheet_view.showGridLines = False

    df_work = data_df.drop(columns=['date_dt'], errors='ignore').copy()

    # Symboles présents dans l'ordre canonique ALL_SYMBOLS
    syms_present = [s for s in ALL_SYMBOLS if s in df_work['symbole'].unique()]
    # Ajouter éventuels symboles hors liste (nouveaux listings)
    extra = [s for s in sorted(df_work['symbole'].unique()) if s not in syms_present]
    syms_present += extra

    # Toutes les dates triées (format DD/MM/YYYY → tri chronologique)
    all_dates = sorted(
        df_work['date'].unique(),
        key=lambda d: (d[6:10], d[3:5], d[0:2])  # YYYY, MM, DD
    )

    # Index rapide : (date, symbole) → ligne
    df_idx = df_work.set_index(['date', 'symbole'])

    # ── Ligne 1 : "Date" + noms de symboles (fusionnés sur N_FIELDS colonnes) ─
    # Col 1 = Date (fusionnée lignes 1-2)
    c = ws1.cell(1, 1, 'Date')
    c.font  = _font(bold=True, color=C_GOLD, size=11, name='Calibri')
    c.fill  = _fill(C_DARK)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = thick_brd
    ws1.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws1.column_dimensions['A'].width = 12
    ws1.row_dimensions[1].height = 22
    ws1.row_dimensions[2].height = 16

    for si, sym in enumerate(syms_present):
        col_start = 2 + si * N_FIELDS          # 1-based
        col_end   = col_start + N_FIELDS - 1

        # Couleur alternée du groupe
        bg = C_SYM_HDR[si % len(C_SYM_HDR)]

        # Cellule symbole (ligne 1, fusionnée sur N_FIELDS colonnes)
        c = ws1.cell(1, col_start, sym)
        c.font  = _font(bold=True, color=C_GOLD, size=10)
        c.fill  = _fill(bg)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thick_brd
        if col_end > col_start:
            ws1.merge_cells(
                start_row=1, start_column=col_start,
                end_row=1,   end_column=col_end
            )

        # Ligne 2 : sous-en-têtes des champs
        for fi, (field, label_f, fmt, w) in enumerate(FIELDS):
            cc = col_start + fi
            c2 = ws1.cell(2, cc, label_f)
            c2.font  = _font(bold=True, color=C_WHITE, size=8)
            c2.fill  = _fill(bg)
            c2.alignment = _center()
            c2.border = thin_brd
            ws1.column_dimensions[get_column_letter(cc)].width = w

    # ── Lignes de données ──────────────────────────────────────────────────────
    for ri, dt in enumerate(all_dates, 3):
        row_fill = _fill(C_EVEN) if ri % 2 == 0 else _fill(C_WHITE)

        # Colonne Date
        c = ws1.cell(ri, 1, dt)
        c.font  = _font(bold=False, color=C_DARK, size=9)
        c.fill  = _fill(C_LIGHT_BG)
        c.alignment = _center()
        c.border = thin_brd

        for si, sym in enumerate(syms_present):
            col_start = 2 + si * N_FIELDS
            try:
                row_data = df_idx.loc[(dt, sym)]
            except KeyError:
                row_data = None

            for fi, (field, label_f, num_fmt, w) in enumerate(FIELDS):
                cc = col_start + fi
                val = None
                if row_data is not None:
                    raw = row_data[field] if field in row_data.index else None
                    if raw is not None and not (isinstance(raw, float) and pd.isna(raw)):
                        val = raw
                c = ws1.cell(ri, cc, val)
                c.font   = _font(color='111111', size=9)
                c.fill   = row_fill
                c.border = thin_brd
                c.alignment = Alignment(horizontal='right', vertical='center')
                if val is not None:
                    c.number_format = num_fmt
                    # Coloration variation jour
                    if field == 'variation_jour_pct' and isinstance(val, (int, float)):
                        if val > 0:
                            c.font = _font(color='1A7A40', size=9, bold=True)
                        elif val < 0:
                            c.font = _font(color='B52626', size=9, bold=True)

    # Figer : 2 lignes d'en-tête + colonne Date
    ws1.freeze_panes = 'B3'

    # ══════════════════════════════════════════════════════════════════════════
    # FEUILLES 2 & 3 — PIVOTS SIMPLES (Clôture / Volume)
    # ══════════════════════════════════════════════════════════════════════════
    def _write_simple_pivot(ws, df_src, value_col, num_fmt='#,##0'):
        ws.sheet_view.showGridLines = False
        syms = [s for s in ALL_SYMBOLS if s in df_src['symbole'].unique()]
        extra_s = [s for s in sorted(df_src['symbole'].unique()) if s not in syms]
        syms = syms + extra_s

        dates_s = sorted(
            df_src['date'].unique(),
            key=lambda d: (d[6:10], d[3:5], d[0:2])
        )

        # Pivot pandas
        piv = df_src.pivot_table(index='date', columns='symbole', values=value_col, aggfunc='last')
        piv = piv.reindex(columns=syms)

        headers = ['Date'] + syms
        # Ligne 1 : en-têtes
        for ci, h in enumerate(headers, 1):
            c = ws.cell(1, ci, h)
            c.font = _font(bold=True, color=C_GOLD if ci == 1 else C_WHITE, size=10)
            c.fill = _fill(C_DARK)
            c.alignment = _center()
            c.border = thin_brd
            ws.column_dimensions[get_column_letter(ci)].width = 12 if ci == 1 else 9

        ws.row_dimensions[1].height = 18

        for ri, dt in enumerate(dates_s, 2):
            row_fill = _fill(C_EVEN) if ri % 2 == 0 else _fill(C_WHITE)
            c = ws.cell(ri, 1, dt)
            c.font = _font(color=C_DARK, size=9)
            c.fill = _fill(C_LIGHT_BG)
            c.alignment = _center()
            c.border = thin_brd

            for ci, sym in enumerate(syms, 2):
                try:
                    val = piv.loc[dt, sym]
                    val = None if pd.isna(val) else val
                except (KeyError, TypeError):
                    val = None
                c = ws.cell(ri, ci, val)
                c.font = _font(color='111111', size=9)
                c.fill = row_fill
                c.border = thin_brd
                c.alignment = Alignment(horizontal='right', vertical='center')
                if val is not None:
                    c.number_format = num_fmt

        ws.freeze_panes = 'B2'

    # ── Feuille 2 : Pivot Ouverture + Clôture (2 valeurs par symbole) ──────────
    ws2 = wb.create_sheet("Pivot Cours")
    ws2.sheet_view.showGridLines = False

    syms_p2 = [s for s in ALL_SYMBOLS if s in df_work["symbole"].unique()]
    syms_p2 += [s for s in sorted(df_work["symbole"].unique()) if s not in syms_p2]

    dates_p2 = sorted(df_work["date"].unique(), key=lambda d: (d[6:10], d[3:5], d[0:2]))

    piv_ouv = df_work.pivot_table(index="date", columns="symbole", values="ouverture", aggfunc="last").reindex(columns=syms_p2)
    piv_clo = df_work.pivot_table(index="date", columns="symbole", values="cloture",   aggfunc="last").reindex(columns=syms_p2)

    N_P2 = 2
    C_P2_HDR = ["1B3A5C", "17324F", "142D47", "112840", "0F2339"]

    ws2.row_dimensions[1].height = 22
    ws2.row_dimensions[2].height = 16
    ws2.column_dimensions["A"].width = 12

    c = ws2.cell(1, 1, "Date")
    c.font      = _font(bold=True, color=C_GOLD, size=11)
    c.fill      = _fill(C_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = thick_brd
    ws2.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    for si, sym in enumerate(syms_p2):
        col_s = 2 + si * N_P2
        bg = C_P2_HDR[si % len(C_P2_HDR)]

        c = ws2.cell(1, col_s, sym)
        c.font      = _font(bold=True, color=C_GOLD, size=10)
        c.fill      = _fill(bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thick_brd
        ws2.merge_cells(start_row=1, start_column=col_s, end_row=1, end_column=col_s + 1)

        for fi, lbl in enumerate(["Ouv.", "Clô."]):
            cc = col_s + fi
            c2 = ws2.cell(2, cc, lbl)
            c2.font      = _font(bold=True, color=C_WHITE, size=8)
            c2.fill      = _fill(bg)
            c2.alignment = _center()
            c2.border    = thin_brd
            ws2.column_dimensions[get_column_letter(cc)].width = 9

    for ri, dt in enumerate(dates_p2, 3):
        row_fill = _fill(C_EVEN) if ri % 2 == 0 else _fill(C_WHITE)

        c = ws2.cell(ri, 1, dt)
        c.font      = _font(color=C_DARK, size=9)
        c.fill      = _fill(C_LIGHT_BG)
        c.alignment = _center()
        c.border    = thin_brd

        for si, sym in enumerate(syms_p2):
            col_s = 2 + si * N_P2
            for fi, piv_tbl in enumerate([piv_ouv, piv_clo]):
                try:
                    val = piv_tbl.loc[dt, sym]
                    val = None if pd.isna(val) else val
                except (KeyError, TypeError):
                    val = None
                cc = col_s + fi
                c = ws2.cell(ri, cc, val)
                c.font      = _font(color="111111", size=9)
                c.fill      = row_fill
                c.border    = thin_brd
                c.alignment = Alignment(horizontal="right", vertical="center")
                if val is not None:
                    c.number_format = "#,##0"

    ws2.freeze_panes = "B3"

    ws3 = wb.create_sheet("Pivot Volumes")
    _write_simple_pivot(ws3, df_work, 'volume', '#,##0')

    # ══════════════════════════════════════════════════════════════════════════
    # FEUILLE 4 — DONNÉES BRUTES (format long)
    # ══════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Données Brutes")
    ws4.sheet_view.showGridLines = False
    raw_cols = ['date', 'symbole', 'titre', 'secteur', 'ouverture', 'cloture',
                'variation_jour_pct', 'volume', 'valeur_fcfa', 'cours_reference',
                'variation_annee_pct', 'dividende_net', 'rendement_net_pct', 'per']
    valid_raw = [c for c in raw_cols if c in df_work.columns]
    raw_headers = [c.replace('_', ' ').title() for c in valid_raw]
    ws4.row_dimensions[1].height = 18
    for ci, h in enumerate(raw_headers, 1):
        c = ws4.cell(1, ci, h)
        c.font = _font(bold=True, color=C_GOLD, size=9)
        c.fill = _fill(C_DARK)
        c.alignment = _center()
        c.border = thin_brd
        ws4.column_dimensions[get_column_letter(ci)].width = max(10, len(h) + 2)
    df_raw_sorted = df_work[valid_raw].sort_values(['date', 'symbole'])
    for ri, row in enumerate(df_raw_sorted.itertuples(index=False), 2):
        row_fill = _fill(C_EVEN) if ri % 2 == 0 else _fill(C_WHITE)
        for ci, v in enumerate(row, 1):
            val = None if (isinstance(v, float) and pd.isna(v)) else v
            c = ws4.cell(ri, ci, val)
            c.font = _font(color='111111', size=9)
            c.fill = row_fill
            c.border = thin_brd
    ws4.freeze_panes = 'A2'

    # ══════════════════════════════════════════════════════════════════════════
    # FEUILLE 5 — STATS PAR TITRE
    # ══════════════════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("Stats par Titre")
    ws5.sheet_view.showGridLines = False
    stat_headers = ['Symbole', 'Titre', 'Secteur', 'Dernier Cours', 'Var. Jour %',
                    'Perf. Période %', 'Vol. Moyen', 'Vol. Max', 'PER', 'Rdt. Net %', 'Séances']
    ws5.row_dimensions[1].height = 18
    for ci, h in enumerate(stat_headers, 1):
        c = ws5.cell(1, ci, h)
        c.font = _font(bold=True, color=C_GOLD, size=9)
        c.fill = _fill(C_DARK)
        c.alignment = _center()
        c.border = thin_brd
        ws5.column_dimensions[get_column_letter(ci)].width = max(11, len(h) + 2)

    stat_rows = []
    for sym in syms_present:
        sd = df_work[df_work['symbole'] == sym].sort_values('date')
        if sd.empty: continue
        f, l = sd.iloc[0], sd.iloc[-1]
        perf_ = (l['cloture'] - f['cloture']) / f['cloture'] * 100 if (f['cloture'] and f['cloture'] != 0) else None
        stat_rows.append((
            sym, str(l.get('titre', ''))[:40], l.get('secteur', ''),
            l['cloture'], l.get('variation_jour_pct'), perf_,
            sd['volume'].mean() if 'volume' in sd else None,
            sd['volume'].max()  if 'volume' in sd else None,
            l.get('per'), l.get('rendement_net_pct'), len(sd)
        ))

    for ri, row in enumerate(stat_rows, 2):
        row_fill = _fill(C_EVEN) if ri % 2 == 0 else _fill(C_WHITE)
        for ci, v in enumerate(row, 1):
            val = None if (isinstance(v, float) and pd.isna(v)) else v
            c = ws5.cell(ri, ci, val)
            c.font = _font(color='111111', size=9)
            c.fill = row_fill
            c.border = thin_brd
            c.alignment = Alignment(
                horizontal='right' if isinstance(val, (int, float)) and val is not None else 'left',
                vertical='center'
            )
            # Formatage numérique selon colonne
            if val is not None and isinstance(val, (int, float)):
                if ci in (4,):   c.number_format = '#,##0'       # Cours
                if ci in (5, 6, 10): c.number_format = '+0.00%;-0.00%;0.00%'
                if ci in (7, 8): c.number_format = '#,##0'
                if ci == 9:      c.number_format = '0.00'
            # Couleur perf
            if ci in (5, 6) and isinstance(val, (int, float)):
                if val > 0:  c.font = _font(color='1A7A40', size=9, bold=True)
                elif val < 0: c.font = _font(color='B52626', size=9, bold=True)
    ws5.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# GESTION DES ACTIONS
# ══════════════════════════════════════════════════════════════════════════════

def acheter_actions(symbole: str, quantite: int):
    """
    Achète des actions basées sur le symbole et la quantité.
    """
    # Logique d'achat (à compléter selon la stratégie)
    st.write(f'Achat de {quantite} actions de {symbole}.')


def generer_predictions(donnees: pd.DataFrame) -> dict:
    """
    Génère des prédictions basées sur les données historiques.
    """
    predictions = {}
    # Exemple de logique de prédiction
    for index, row in donnees.iterrows():
        symbole = row['symbole']
        prix_historique = row['prix']
        # Logique de prédiction (à compléter)
        prediction = prix_historique * 1.05  # Exemple : augmentation de 5%
        predictions[symbole] = prediction
    return predictions


def recommandations_investisseur(rendement: float) -> str:
    """
    Fournit des recommandations basées sur le rendement.
    """
    if rendement > 10:
        return "Recommandation : Investir davantage dans cet actif."
    elif rendement > 5:
        return "Recommandation : Maintenir l'investissement actuel."
    else:
        return "Recommandation : Envisager de vendre cet actif."


def calculer_rendement(prix_achat: float, prix_vente: float) -> float:
    """
    Calcule le rendement d'un actif basé sur le prix d'achat et le prix de vente.
    """
    if prix_achat == 0:
        return 0.0
    return (prix_vente - prix_achat) / prix_achat * 100


def vendre_actions(symbole: str, quantite: int):
    """
    Vends des actions basées sur le symbole et la quantité.
    """
    # Logique de vente (à compléter selon la stratégie)
    st.write(f'Vente de {quantite} actions de {symbole}.')

def analyser_recommandations(donnees: pd.DataFrame, horizon: int = 20) -> pd.DataFrame:
    """Produit des signaux explicables depuis les cours de clôture disponibles.

    La projection utilise la tendance linéaire récente, limitée à ±20 %. C'est
    un indicateur de recherche : elle n'est ni un conseil financier personnalisé,
    ni une garantie de performance.
    """
    resultats = []
    for symbole, serie in donnees.groupby('symbole'):
        serie = serie.sort_values('date_dt').dropna(subset=['cloture']).copy()
        cours = pd.to_numeric(serie['cloture'], errors='coerce').dropna()
        if len(cours) < 10 or cours.iloc[-1] <= 0:
            continue

        rendements = cours.pct_change().dropna()
        fenetre = min(20, len(cours))
        recents = cours.iloc[-fenetre:].reset_index(drop=True)
        x = pd.Series(range(len(recents)), dtype=float)
        denominateur = ((x - x.mean()) ** 2).sum()
        pente = float(((x - x.mean()) * (recents - recents.mean())).sum() / denominateur) if denominateur else 0.0
        cours_actuel = float(cours.iloc[-1])
        tendance_pct = pente / cours_actuel * 100
        rendement_20j = ((cours_actuel / float(cours.iloc[-fenetre])) - 1) * 100
        volatilite = float(rendements.tail(fenetre).std() * 100) if len(rendements) > 1 else 0.0
        projection_pct = max(-20.0, min(20.0, tendance_pct * horizon))
        prix_projete = cours_actuel * (1 + projection_pct / 100)
        score = (0.45 * rendement_20j) + (0.45 * projection_pct) - (0.10 * volatilite)
        signal = ('ACHAT À ÉTUDIER' if score >= 4 and projection_pct > 0 else
                  'VENTE À ÉTUDIER' if score <= -4 and projection_pct < 0 else
                  'CONSERVER / SURVEILLER')
        dernier = serie.iloc[-1]
        resultats.append({
            'Symbole': symbole, 'Titre': str(dernier.get('titre', ''))[:35],
            'Cours actuel (FCFA)': round(cours_actuel, 0),
            'Rendement 20 séances (%)': round(rendement_20j, 2),
            'Tendance / séance (%)': round(tendance_pct, 3),
            f'Projection {horizon} séances (FCFA)': round(prix_projete, 0),
            f'Projection {horizon} séances (%)': round(projection_pct, 2),
            'Volatilité (%)': round(volatilite, 2), 'Score': round(score, 2),
            'Signal': signal, 'Observations': len(cours),
        })
    return pd.DataFrame(resultats)


if 'cached_df' not in st.session_state:
    st.session_state.cached_df = load_cache()
if 'extraction_log' not in st.session_state:
    st.session_state.extraction_log = []
if 'is_extracting' not in st.session_state:
    st.session_state.is_extracting = False
if 'extract_log_dict' not in st.session_state:
    st.session_state.extract_log_dict = load_log()
if 'iv_trigger_excel' not in st.session_state:
    st.session_state.iv_trigger_excel = False


# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════


def _run_extraction(codes: list[str], label: str = "", delay: float = 0.4):
    """Lance l'extraction pour une liste de codes de dates."""
                                                                                                                                                                                                                                    # Exemple de logique pour acheter ou vendre des actions
    donnees_historique = pd.DataFrame()  # Remplacer par les données réelles
    predictions = generer_predictions(donnees_historique)
    for code in []:  # ancien exemple désactivé
        # Logique d'extraction des données
        # ...
        symbole = predictions[code]
        symbole = "AAPL"  # Remplacer par le symbole réel
        quantite = 10  # Remplacer par la quantité réelle
        prix_achat = 100.0  # Remplacer par le prix d'achat réel
        prix_vente = 110.0  # Remplacer par le prix de vente réel
        rendement = calculer_rendement(prix_achat, prix_vente)
        prediction = rendement > 5  # Exemple de condition de prédiction
        # Supposons que nous avons une prédiction pour le symbole et la quantité
        # rendement = calculer_rendement(prix_achat, prix_vente)
        st.write(recommandations_investisseur(rendement))
        if prediction:
            acheter_actions(symbole, quantite)
        else:
            vendre_actions(symbole, quantite)
    if not codes:
        st.sidebar.success("Rien à extraire.")
        return

    log = st.session_state.extract_log_dict
    total = len(codes)
    all_new_rows = []
    live_log = []

    # UI inline dans le sidebar
    prog_title = st.sidebar.empty()
    prog_bar   = st.sidebar.progress(0)
    prog_status = st.sidebar.empty()
    prog_log    = st.sidebar.empty()

    prog_title.markdown(f"**⏳ Extraction : {label}**")
    n_ok = n_fail = n_skip = 0

    for i, code in enumerate(codes):
        disp = code_to_display(code)
        prog_status.markdown(f"<small style='color:#7A8699'>📥 {disp} ({i+1}/{total})</small>", unsafe_allow_html=True)

        pdf_bytes = download_pdf(code)

        if pdf_bytes:
            rows = parse_actions(pdf_bytes, disp)
            if rows:
                all_new_rows.extend(rows)
                log[code] = {'status': 'ok', 'rows': len(rows), 'ts': datetime.now().isoformat()}
                live_log.append(f'<div class="log-ok">✓ {disp} — {len(rows)} titres</div>')
                n_ok += 1
            else:
                log[code] = {'status': 'ok_empty', 'rows': 0, 'ts': datetime.now().isoformat()}
                live_log.append(f'<div class="log-warn">⚠ {disp} — PDF vide</div>')
                n_skip += 1
        else:
            log[code] = {'status': 'not_found', 'ts': datetime.now().isoformat()}
            live_log.append(f'<div class="log-err">✗ {disp} — non disponible</div>')
            n_fail += 1

        prog_bar.progress((i + 1) / total)

        # Afficher les 12 dernières lignes de log
        recent = live_log[-12:]
        prog_log.markdown(
            f'<div class="progress-block">{"".join(recent)}</div>',
            unsafe_allow_html=True
        )

        time.sleep(delay)

    # Sauvegarder
    if all_new_rows:
        updated = merge_into_cache(st.session_state.cached_df, all_new_rows)
        save_cache(updated)
        st.session_state.cached_df = updated

    save_log(log)
    st.session_state.extract_log_dict = log

    # Résumé final
    prog_status.empty()
    prog_title.markdown(f"""
    <div class="brvm-card" style="padding:10px 14px;">
        <div style="color:var(--gold);font-family:'Space Mono',monospace;font-size:.82rem;margin-bottom:6px;">
            ✅ Extraction terminée — {label}
        </div>
        <div style="font-size:.75rem;color:#7A8699;">
            <span class="dot dot-g"></span>{n_ok} séances · 
            <span class="dot dot-o"></span>{n_skip} vides · 
            <span class="dot dot-r"></span>{n_fail} introuvables<br>
            <strong style="color:var(--text);">{len(all_new_rows):,}</strong> nouvelles lignes ajoutées
        </div>
    </div>
    """, unsafe_allow_html=True)
    time.sleep(1)
    st.rerun()



with st.sidebar:
    st.markdown("""
    <div style="text-align:center;padding:10px 0 18px;">
        <div style="font-family:'Space Mono',monospace;color:#C9A84C;font-size:1.5rem;font-weight:700;letter-spacing:.12em;">
            BRVM
        </div>
        <div style="color:#7A8699;font-size:.68rem;letter-spacing:.18em;text-transform:uppercase;margin-top:3px;">
            Data Extractor v3
        </div>
        <div style="margin-top:10px;">
            <span class="badge badge-info">47 Sociétés · UEMOA · FCFA</span>
        </div>
    </div>
    <hr class="divider">
    """, unsafe_allow_html=True)

    # ── Statut du cache ──────────────────────────────────────────────
    cached_df = st.session_state.cached_df
    if cached_df is not None and not cached_df.empty:
        n_sessions = cached_df['date'].nunique()
        n_records  = len(cached_df)
        n_symbols  = cached_df['symbole'].nunique()
        min_d = pd.to_datetime(cached_df['date'], format='%d/%m/%Y', errors='coerce').min()
        max_d = pd.to_datetime(cached_df['date'], format='%d/%m/%Y', errors='coerce').max()
        st.markdown(f"""
        <div class="brvm-card" style="padding:12px 14px;margin-bottom:10px;">
            <div style="color:#7A8699;font-size:.68rem;text-transform:uppercase;letter-spacing:.08em;margin-bottom:8px;">
                <span class="dot dot-g"></span>Cache actif
            </div>
            <div style="display:grid;grid-template-columns:1fr 1fr;gap:6px;">
                <div style="background:var(--surface3);border-radius:6px;padding:7px;text-align:center;">
                    <div style="color:#7A8699;font-size:.64rem;">Séances</div>
                    <div style="color:var(--gold);font-family:'Space Mono',monospace;font-size:1rem;">{n_sessions}</div>
                </div>
                <div style="background:var(--surface3);border-radius:6px;padding:7px;text-align:center;">
                    <div style="color:#7A8699;font-size:.64rem;">Titres</div>
                    <div style="color:var(--gold);font-family:'Space Mono',monospace;font-size:1rem;">{n_symbols}</div>
                </div>
                <div style="background:var(--surface3);border-radius:6px;padding:7px;text-align:center;">
                    <div style="color:#7A8699;font-size:.64rem;">Lignes</div>
                    <div style="color:var(--text);font-family:'Space Mono',monospace;font-size:.85rem;">{n_records:,}</div>
                </div>
                <div style="background:var(--surface3);border-radius:6px;padding:7px;text-align:center;">
                    <div style="color:#7A8699;font-size:.64rem;">Années</div>
                    <div style="color:var(--text);font-family:'Space Mono',monospace;font-size:.85rem;">
                        {min_d.year if pd.notna(min_d) else '?'}–{max_d.year if pd.notna(max_d) else '?'}
                    </div>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)
    else:
        st.info("📭 Aucune donnée en cache")

    st.markdown("---")

    # ── Mode d'extraction ──────────────────────────────────────────────
    st.markdown("### ⚙️ Mode d'extraction")
    mode = st.radio(
        "Choisir le mode",
        ["📅 Bulletins récents", "📆 Mois spécifique", "📚 Année complète",
         "🗓️ Intervalle de dates", "📂 Import Excel"],
        label_visibility="collapsed"
    )

    st.markdown("---")

    # ══════════════════════════════════════════════════════════════════
    # MODE 1 : Bulletins récents
    # ══════════════════════════════════════════════════════════════════
    if mode == "📅 Bulletins récents":
        st.markdown("#### 📅 Bulletins récents")
        st.caption("Scrape la liste officielle BRVM")

        with st.spinner("Chargement liste BRVM..."):
            recent_codes = fetch_recent_boc_list()

        if recent_codes:
            st.caption(f"**{len(recent_codes)}** bulletins disponibles")
            n_dl = st.slider("Nombre à extraire", 1, min(len(recent_codes),60), 5)
            to_dl = recent_codes[:n_dl]

            # Filtrer déjà extraits
            log = st.session_state.extract_log_dict
            already = {c for c,v in log.items() if v.get('status')=='ok'}
            new_only = [c for c in to_dl if c not in already]

            if new_only:
                st.info(f"🆕 {len(new_only)} nouveaux · {len(to_dl)-len(new_only)} déjà extraits")
            else:
                st.success("✅ Tous déjà extraits")

            if st.button("🚀 EXTRAIRE", use_container_width=True, key="btn_recent"):
                _run_extraction(new_only, label="Bulletins récents")
        else:
            st.warning("Site BRVM inaccessible ou réseau restreint.")
            st.caption("👉 Utilisez 'Année complète' ou 'Import Excel'")

    # ══════════════════════════════════════════════════════════════════
    # MODE 2 : Mois spécifique
    # ══════════════════════════════════════════════════════════════════
    elif mode == "📆 Mois spécifique":
        st.markdown("#### 📆 Mois spécifique")
        col_y, col_m = st.columns(2)
        with col_y:
            year_sel = st.selectbox("Année", list(range(CURRENT_YEAR, FIRST_YEAR-1, -1)), key="yr_month")
        with col_m:
            month_sel = st.selectbox("Mois",
                [(i, datetime(2000,i,1).strftime('%b')) for i in range(1,13)],
                format_func=lambda x: x[1], key="mo_sel")
            month_num = month_sel[0]

        codes = get_candidate_dates_range(year_sel, month_num, month_num)
        log = st.session_state.extract_log_dict
        already = {c for c,v in log.items() if v.get('status')=='ok'}
        new_codes = [c for c in codes if c not in already]

        st.caption(f"~{len(codes)} séances candidates · {len(new_codes)} à télécharger")

        if st.button("🚀 EXTRAIRE CE MOIS", use_container_width=True, key="btn_month"):
            _run_extraction(new_codes, label=f"{datetime(year_sel,month_num,1).strftime('%B %Y')}")

    # ══════════════════════════════════════════════════════════════════
    # MODE 3 : Année complète
    # ══════════════════════════════════════════════════════════════════
    elif mode == "📚 Année complète":
        st.markdown("#### 📚 Année complète")
        st.caption("Télécharge & extrait tous les BOC d'une année")

        year_full = st.selectbox(
            "Sélectionner l'année",
            list(range(CURRENT_YEAR, FIRST_YEAR-1, -1)),
            key="yr_full"
        )

        all_codes = get_candidate_dates(year_full)
        log = st.session_state.extract_log_dict
        already_ok   = {c for c,v in log.items() if v.get('status')=='ok'}
        already_fail = {c for c,v in log.items() if v.get('status')=='not_found'}
        new_codes    = [c for c in all_codes if c not in already_ok and c not in already_fail]
        done_codes   = [c for c in all_codes if c in already_ok]

        # Stats visuelles
        pct_done = len(done_codes)/len(all_codes)*100 if all_codes else 0
        st.markdown(f"""
        <div class="brvm-card" style="padding:12px;">
            <div style="display:flex;justify-content:space-between;margin-bottom:8px;">
                <span style="color:#7A8699;font-size:.75rem;">{year_full} — Progression</span>
                <span style="color:var(--gold);font-family:'Space Mono',monospace;font-size:.8rem;">{pct_done:.0f}%</span>
            </div>
            <div style="background:var(--surface);border-radius:4px;height:6px;overflow:hidden;">
                <div style="background:linear-gradient(90deg,#C9A84C,#E8C96B);width:{pct_done:.1f}%;height:100%;border-radius:4px;transition:width .3s;"></div>
            </div>
            <div style="display:flex;gap:12px;margin-top:8px;font-size:.7rem;color:#7A8699;">
                <span><span class="dot dot-g"></span>{len(done_codes)} extraits</span>
                <span><span class="dot dot-o"></span>{len(new_codes)} restants</span>
                <span><span class="dot dot-r"></span>{len(already_fail)} introuvables</span>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # Options avancées
        with st.expander("⚙️ Options avancées"):
            skip_known_fail = st.checkbox("Ignorer les dates déjà échouées", value=True)
            if not skip_known_fail:
                new_codes = [c for c in all_codes if c not in already_ok]
            
            concurrent = st.slider("Connexions simultanées", 1, 5, 3,
                                    help="Plus = plus rapide mais plus de risque de blocage")
            delay = st.slider("Délai entre requêtes (sec)", 0.1, 2.0, 0.5, 0.1)

        if new_codes:
            # Estimation du temps
            est_min = len(new_codes) * delay / 60
            est_max = len(new_codes) * (delay + 2) / 60
            st.caption(f"⏱️ Estimation : {est_min:.0f}–{est_max:.0f} min pour {len(new_codes)} dates")

            if st.button(f"🚀 EXTRAIRE {year_full}", use_container_width=True, key="btn_year"):
                _run_extraction(new_codes, label=str(year_full), delay=delay)
        else:
            st.success(f"✅ {year_full} entièrement extrait !")

        # Bouton reset année
        if done_codes and st.button(f"🔄 Ré-extraire {year_full}", use_container_width=True, key="btn_reextract"):
            log = st.session_state.extract_log_dict
            for c in all_codes:
                log.pop(c, None)
            save_log(log)
            st.session_state.extract_log_dict = log
            # Remove year data from cache
            if st.session_state.cached_df is not None:
                prefix = str(year_full)
                mask = ~st.session_state.cached_df['date'].str.endswith(prefix)
                # date format DD/MM/YYYY, year is last 4
                mask = ~st.session_state.cached_df['date'].str.contains(f'/{year_full}')
                st.session_state.cached_df = st.session_state.cached_df[~mask] if (~mask).any() else st.session_state.cached_df
            st.rerun()

    # ══════════════════════════════════════════════════════════════════
    # MODE 5 : Intervalle de dates (multi-années)
    # ══════════════════════════════════════════════════════════════════
    elif mode == "🗓️ Intervalle de dates":
        st.markdown("#### 🗓️ Intervalle de dates")
        st.caption("Télécharge tous les BOC entre deux dates (multi-années)")

        today = datetime.now().date()
        first_date = date(FIRST_YEAR, 1, 1)

        col_d1, col_d2 = st.columns(2)
        with col_d1:
            dt_start = st.date_input(
                "Date de début",
                value=date(today.year - 1, 1, 1),
                min_value=first_date,
                max_value=today,
                key="iv_start"
            )
        with col_d2:
            dt_end = st.date_input(
                "Date de fin",
                value=today,
                min_value=first_date,
                max_value=today,
                key="iv_end"
            )

        if dt_start > dt_end:
            st.error("⚠️ La date de début doit être antérieure à la date de fin.")
        else:
            all_codes = get_candidate_dates_interval(dt_start, dt_end)
            log = st.session_state.extract_log_dict
            already_ok   = {c for c, v in log.items() if v.get('status') == 'ok'}
            already_fail = {c for c, v in log.items() if v.get('status') == 'not_found'}
            new_codes    = [c for c in all_codes if c not in already_ok and c not in already_fail]
            done_codes   = [c for c in all_codes if c in already_ok]

            n_years = dt_end.year - dt_start.year + 1
            pct_done = len(done_codes) / len(all_codes) * 100 if all_codes else 0

            st.markdown(f"""
            <div class="brvm-card" style="padding:12px;">
                <div style="display:flex;justify-content:space-between;margin-bottom:8px;">
                    <span style="color:#7A8699;font-size:.75rem;">
                        {dt_start.strftime('%d/%m/%Y')} → {dt_end.strftime('%d/%m/%Y')}
                        &nbsp;({n_years} an{'s' if n_years > 1 else ''})
                    </span>
                    <span style="color:var(--gold);font-family:'Space Mono',monospace;font-size:.8rem;">{pct_done:.0f}%</span>
                </div>
                <div style="background:var(--surface);border-radius:4px;height:6px;overflow:hidden;">
                    <div style="background:linear-gradient(90deg,#C9A84C,#E8C96B);width:{pct_done:.1f}%;height:100%;border-radius:4px;transition:width .3s;"></div>
                </div>
                <div style="display:flex;gap:12px;margin-top:8px;font-size:.7rem;color:#7A8699;">
                    <span><span class="dot dot-g"></span>{len(all_codes)} séances candidates</span>
                    <span><span class="dot dot-b"></span>{len(done_codes)} extraites</span>
                    <span><span class="dot dot-o"></span>{len(new_codes)} restantes</span>
                    <span><span class="dot dot-r"></span>{len(already_fail)} introuvables</span>
                </div>
            </div>
            """, unsafe_allow_html=True)

            with st.expander("⚙️ Options avancées"):
                skip_iv_fail = st.checkbox("Ignorer les dates déjà échouées", value=True, key="iv_skip")
                if not skip_iv_fail:
                    new_codes = [c for c in all_codes if c not in already_ok]
                iv_delay = st.slider("Délai entre requêtes (sec)", 0.1, 2.0, 0.4, 0.1, key="iv_delay")
                iv_export_auto = st.checkbox(
                    "Export Excel automatique après extraction",
                    value=True, key="iv_auto_xl",
                    help="Génère et propose le téléchargement Excel dès la fin de l'extraction"
                )

            if new_codes:
                est_min = len(new_codes) * iv_delay / 60
                est_max = len(new_codes) * (iv_delay + 2) / 60
                st.caption(f"⏱️ Estimation : {est_min:.0f}–{est_max:.0f} min pour {len(new_codes)} dates")
            else:
                st.success("✅ Toutes les dates de cet intervalle sont déjà extraites !")

            # ── Bouton principal extraction ────────────────────────────
            label_iv = f"{dt_start.strftime('%d/%m/%Y')} → {dt_end.strftime('%d/%m/%Y')}"

            col_b1, col_b2 = st.columns(2)
            with col_b1:
                if new_codes:
                    if st.button("🚀 EXTRAIRE INTERVALLE", use_container_width=True, key="btn_iv"):
                        _run_extraction(new_codes, label=label_iv, delay=iv_delay)
                        # After extraction, trigger Excel if auto-export enabled
                        if iv_export_auto:
                            st.session_state['iv_trigger_excel'] = True

            with col_b2:
                # Export Excel direct sur les données déjà en cache pour cet intervalle
                if st.button("📗 EXPORTER EXCEL", use_container_width=True, key="btn_iv_xl"):
                    st.session_state['iv_trigger_excel'] = True

            # ── Génération / téléchargement Excel ─────────────────────
            if st.session_state.get('iv_trigger_excel') and st.session_state.cached_df is not None:
                df_iv = st.session_state.cached_df.copy()
                # Filter to interval
                try:
                    df_iv['_dt'] = pd.to_datetime(df_iv['date'], format='%d/%m/%Y', errors='coerce')
                    df_iv = df_iv[(df_iv['_dt'].dt.date >= dt_start) & (df_iv['_dt'].dt.date <= dt_end)]
                    df_iv = df_iv.drop(columns=['_dt'])
                except Exception:
                    pass

                if df_iv.empty:
                    st.warning("⚠️ Aucune donnée dans cet intervalle. Lancez l'extraction d'abord.")
                else:
                    with st.spinner("📗 Génération Excel en cours…"):
                        xl_bytes = generate_excel_workbook(df_iv, label=label_iv)

                    fname = (
                        f"BRVM_{dt_start.strftime('%Y%m%d')}_"
                        f"{dt_end.strftime('%Y%m%d')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
                    )
                    st.markdown(f"""
                    <div class="brvm-card" style="padding:10px 14px;margin-top:8px;">
                        <div style="color:var(--gold);font-family:'Space Mono',monospace;font-size:.8rem;margin-bottom:6px;">
                            📗 Excel prêt — {label_iv}
                        </div>
                        <div style="font-size:.72rem;color:#7A8699;">
                            {len(df_iv):,} lignes · {df_iv['symbole'].nunique()} titres ·
                            {df_iv['date'].nunique()} séances · 4 feuilles
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
                    st.download_button(
                        "⬇️ Télécharger Excel",
                        data=xl_bytes,
                        file_name=fname,
                        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        use_container_width=True,
                        key="dl_iv_xl"
                    )
                    st.session_state['iv_trigger_excel'] = False

    # ══════════════════════════════════════════════════════════════════
    # MODE 4 : Import Excel
    # ══════════════════════════════════════════════════════════════════
    elif mode == "📂 Import Excel":
        st.markdown("#### 📂 Import fichier Excel")
        uploaded = st.file_uploader("Fichier Excel BRVM", type=['xlsx'], label_visibility="collapsed")
        if uploaded:
            try:
                xl = pd.ExcelFile(uploaded)
                sheet = "Prix et Volumes" if "Prix et Volumes" in xl.sheet_names else xl.sheet_names[0]
                df_imp = xl.parse(sheet)
                col_map = {
                    'Date':'date','Symbole':'symbole','Cours Ouverture':'ouverture',
                    'Cours Clôture':'cloture','Cours Précédent':'cours_reference',
                    'Variation Jour (%)':'variation_jour_pct','Volume':'volume',
                    'Valeur Transigée (FCFA)':'valeur_fcfa'
                }
                df_imp = df_imp.rename(columns=col_map)
                df_imp['date'] = df_imp['date'].astype(str)
                for c in ['secteur_code','secteur','titre','variation_annee_pct',
                          'dividende_net','date_dividende','rendement_net_pct','per']:
                    if c not in df_imp.columns: df_imp[c] = None
                updated = merge_into_cache(st.session_state.cached_df, df_imp.to_dict('records'))
                save_cache(updated)
                st.session_state.cached_df = updated
                st.success(f"✅ {len(df_imp)} lignes importées depuis {sheet}")
                st.rerun()
            except Exception as e:
                st.error(f"Erreur : {e}")

    # ── Gestion du cache ──────────────────────────────────────────────
    st.markdown("---")
    with st.expander("🗄️ Gestion du cache"):
        if st.session_state.cached_df is not None and not st.session_state.cached_df.empty:
            csv_cache = st.session_state.cached_df.to_csv(index=False).encode('utf-8-sig')
            st.download_button("⬇️ Sauvegarder cache CSV", csv_cache,
                               f"brvm_cache_{datetime.now().strftime('%Y%m%d')}.csv", "text/csv",
                               use_container_width=True)
        if st.button("🗑️ Vider tout le cache", use_container_width=True, key="btn_clear"):
            if CACHE_FILE.exists(): CACHE_FILE.unlink()
            if LOG_FILE.exists(): LOG_FILE.unlink()
            st.session_state.cached_df = None
            st.session_state.extract_log_dict = {}
            st.rerun()

        # Vider une année spécifique
        year_del = st.selectbox("Supprimer une année", [None]+list(range(CURRENT_YEAR, FIRST_YEAR-1,-1)), key="yr_del")
        if year_del and st.button(f"🗑️ Supprimer {year_del}", use_container_width=True, key="btn_del_yr"):
            df_c = st.session_state.cached_df
            if df_c is not None:
                mask = df_c['date'].str.endswith(str(year_del))
                st.session_state.cached_df = df_c[~mask].reset_index(drop=True)
                save_cache(st.session_state.cached_df)
            log = st.session_state.extract_log_dict
            to_del = [k for k in log if k[:4]==str(year_del)]
            for k in to_del: del log[k]
            save_log(log); st.session_state.extract_log_dict = log
            st.success(f"✅ Données {year_del} supprimées"); st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# FONCTION D'EXTRACTION (définie après le sidebar pour accès à session_state)
# ══════════════════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════════════════
# MAIN — CONTENU PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════

# En-tête
st.markdown("""
<div class="brvm-header">
    <div>
        <h1 style="margin:0;">📈 BRVM Data Extractor v3</h1>
        <p style="color:#7A8699;margin:4px 0 0;font-size:.82rem;">
            Extraction historique automatisée — Bulletins Officiels de la Cote · UEMOA · Intervalle libre
        </p>
    </div>
    <div style="margin-left:auto;text-align:right;">
        <span class="badge badge-info">bfin.brvm.org</span>
    </div>
</div>
""", unsafe_allow_html=True)

# ── Pas de données ──────────────────────────────────────────────────────────
cached_df = st.session_state.cached_df
if cached_df is None or cached_df.empty:
    st.markdown("""
    <div class="brvm-card" style="text-align:center;padding:60px 40px;">
        <div style="font-size:3.5rem;margin-bottom:14px;">📊</div>
        <h2 style="color:#C9A84C;margin-bottom:10px;">Prêt à extraire</h2>
        <p style="color:#7A8699;max-width:460px;margin:0 auto 20px;font-size:.9rem;">
            Sélectionnez un mode d'extraction dans le panneau gauche :<br>
            <b>Bulletins récents</b> pour les dernières séances, 
            <b>Année complète</b> pour l'historique long terme.
        </p>
        <div style="display:flex;gap:10px;justify-content:center;flex-wrap:wrap;">
            <span class="badge badge-info">📅 Bulletins récents</span>
            <span class="badge badge-flat">📆 Mois spécifique</span>
            <span class="badge badge-info">📚 Année complète (2015–2026)</span>
            <span class="badge badge-up">🗓️ Intervalle de dates ⭐ Nouveau</span>
            <span class="badge badge-flat">📂 Import Excel</span>
        </div>
    </div>
    """, unsafe_allow_html=True)
    st.stop()


# ── Préparation données ─────────────────────────────────────────────────────
df = cached_df.copy()
df['date_dt'] = pd.to_datetime(df['date'], format='%d/%m/%Y', errors='coerce')
df = df.dropna(subset=['date_dt']).sort_values('date_dt')

latest_dt  = df['date_dt'].max()
latest_df  = df[df['date_dt'] == latest_dt]
all_dates  = sorted(df['date_dt'].unique(), reverse=True)
all_years  = sorted(df['date_dt'].dt.year.unique(), reverse=True)
avail_syms = sorted(df['symbole'].dropna().unique())


# ── KPIs ────────────────────────────────────────────────────────────────────
total_vol_last = latest_df['volume'].sum()
total_val_last = latest_df['valeur_fcfa'].sum()
avg_var_last   = latest_df['variation_jour_pct'].mean()
n_sessions     = len(all_dates)

c1,c2,c3,c4,c5,c6 = st.columns(6)
with c1: st.metric("Séances", f"{n_sessions:,}")
with c2: st.metric("Années couvertes", f"{len(all_years)}")
with c3: st.metric("Titres", f"{len(avail_syms)}")
with c4: st.metric("Enregistrements", f"{len(df):,}")
with c5: st.metric("Volume (dernière séance)", fmt_num(total_vol_last))
with c6: st.metric("Variation moy.", f"{avg_var_last:+.2f}%" if pd.notna(avg_var_last) else "—",
                   delta=f"{avg_var_last:+.2f}%" if pd.notna(avg_var_last) else None)

st.markdown("<br>", unsafe_allow_html=True)

# ── ONGLETS ─────────────────────────────────────────────────────────────────
tab_dash, tab_seance, tab_titre, tab_ts, tab_histo, tab_export = st.tabs([
    "🏠 Tableau de bord",
    "📅 Séance",
    "🔍 Par titre",
    "📈 Séries temporelles",
    "🗓️ Historique annuel",
    "💾 Export",
])


# ════════════════════════════════════════════════════════════════════════════
# ONGLET 1 — TABLEAU DE BORD
# ════════════════════════════════════════════════════════════════════════════
with tab_dash:
    col_l, col_r = st.columns([3, 2])

    with col_l:
        st.markdown(f"### Dernière séance : {latest_dt.strftime('%d/%m/%Y')}")
        st.caption(f"{len(latest_df)} titres cotés")

        if not latest_df.empty:
            # Hausse / Baisse
            top5_up   = latest_df[latest_df['variation_jour_pct']>0].nlargest(5,'variation_jour_pct')
            top5_down = latest_df[latest_df['variation_jour_pct']<0].nsmallest(5,'variation_jour_pct')
            ch, cd = st.columns(2)
            with ch:
                st.markdown("#### 🟢 Meilleures hausses")
                for _, r in top5_up.iterrows():
                    st.markdown(f"""
                    <div class="ticker-row">
                        <div>
                            <span style="font-family:'Space Mono',monospace;color:#DDE3EF;font-size:.85rem;font-weight:700;">{r['symbole']}</span>
                            <span style="color:#7A8699;font-size:.7rem;margin-left:6px;">{str(r.get('titre',''))[:20]}</span>
                        </div>
                        <div style="text-align:right;">
                            <span style="color:#7A8699;font-size:.78rem;">{r['cloture']:,.0f} </span>
                            {badge(r['variation_jour_pct'])}
                        </div>
                    </div>""", unsafe_allow_html=True)
            with cd:
                st.markdown("#### 🔴 Plus fortes baisses")
                for _, r in top5_down.iterrows():
                    st.markdown(f"""
                    <div class="ticker-row">
                        <div>
                            <span style="font-family:'Space Mono',monospace;color:#DDE3EF;font-size:.85rem;font-weight:700;">{r['symbole']}</span>
                            <span style="color:#7A8699;font-size:.7rem;margin-left:6px;">{str(r.get('titre',''))[:20]}</span>
                        </div>
                        <div style="text-align:right;">
                            <span style="color:#7A8699;font-size:.78rem;">{r['cloture']:,.0f} </span>
                            {badge(r['variation_jour_pct'])}
                        </div>
                    </div>""", unsafe_allow_html=True)

        # Graphique secteurs
        st.markdown("<hr class='divider'>", unsafe_allow_html=True)
        st.markdown("#### 📊 Valeur par secteur — dernière séance")
        if 'secteur' in latest_df.columns:
            sv = latest_df.groupby('secteur').agg(valeur=('valeur_fcfa','sum')).reset_index().sort_values('valeur')
            try:
                import plotly.graph_objects as go
                fig = go.Figure(go.Bar(
                    x=sv['valeur']/1e6, y=sv['secteur'], orientation='h',
                    marker=dict(color=sv['valeur'], colorscale=[[0,'#172035'],[.5,'#9A6F1E'],[1,'#C9A84C']],
                                line=dict(color='rgba(201,168,76,.25)',width=1)),
                    text=[fmt_num(v,' M') for v in sv['valeur']], textposition='outside',
                    textfont=dict(color='#7A8699',size=10)
                ))
                fig.update_layout(
                    paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                    xaxis=dict(title='M FCFA', color='#7A8699', gridcolor='rgba(201,168,76,.07)'),
                    yaxis=dict(color='#DDE3EF', tickfont=dict(size=11)),
                    margin=dict(l=0,r=90,t=10,b=30), height=260, font=dict(family='DM Sans')
                )
                st.plotly_chart(fig, use_container_width=True)
            except ImportError:
                st.dataframe(sv, hide_index=True, use_container_width=True)

    with col_r:
        st.markdown("### 📌 Statistiques marché")
        nb_h  = (latest_df['variation_jour_pct']>0).sum()
        nb_b  = (latest_df['variation_jour_pct']<0).sum()
        nb_0  = (latest_df['variation_jour_pct']==0).sum()
        pct_h = nb_h/len(latest_df)*100 if len(latest_df) else 0
        pct_b = nb_b/len(latest_df)*100 if len(latest_df) else 0
        st.markdown(f"""
        <div class="brvm-card">
            <div style="color:#7A8699;font-size:.7rem;text-transform:uppercase;letter-spacing:.08em;margin-bottom:10px;">Sentiment</div>
            <div style="margin-bottom:6px;"><span class="dot dot-g"></span><span style="color:#2ECC71">{nb_h} titres en hausse ({pct_h:.0f}%)</span></div>
            <div style="margin-bottom:6px;"><span class="dot dot-r"></span><span style="color:#E74C3C">{nb_b} titres en baisse ({pct_b:.0f}%)</span></div>
            <div style="margin-bottom:10px;"><span class="dot dot-o"></span><span style="color:#7A8699">{nb_0} inchangés</span></div>
            <div style="background:var(--surface);border-radius:4px;height:7px;overflow:hidden;">
                <div style="background:linear-gradient(90deg,#2ECC71 {pct_h:.0f}%,#E74C3C {pct_h:.0f}%);height:100%;"></div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("**🏆 Top 5 volumes**")
        for _, r in latest_df.nlargest(5,'volume').iterrows():
            st.markdown(f"""
            <div class="ticker-row">
                <span style="font-family:'Space Mono',monospace;color:#C9A84C;font-size:.82rem;">{r['symbole']}</span>
                <div>
                    <div style="color:#DDE3EF;font-size:.78rem;text-align:right;">{fmt_num(r['volume'])}</div>
                    <div style="color:#7A8699;font-size:.68rem;text-align:right;">{r['cloture']:,.0f} FCFA</div>
                </div>
            </div>""", unsafe_allow_html=True)

        # Évolution du nombre de séances par mois
        st.markdown("<hr class='divider'>", unsafe_allow_html=True)
        st.markdown("**📅 Séances par mois**")
        monthly = df.groupby(df['date_dt'].dt.to_period('M')).agg(
            séances=('date_dt','nunique')).reset_index()
        monthly['date_dt'] = monthly['date_dt'].astype(str)
        monthly = monthly.sort_values('date_dt', ascending=False).head(12)
        monthly.columns = ['Mois','Séances']
        st.dataframe(monthly, hide_index=True, use_container_width=True, height=280)


# ════════════════════════════════════════════════════════════════════════════
# ONGLET 2 — SÉANCE
# ════════════════════════════════════════════════════════════════════════════
with tab_dash:
    st.markdown("### Recommandations quantitatives")
    st.caption("Signaux calculés à partir des rendements, de la tendance récente et de la volatilité. À utiliser comme outil d'analyse, pas comme conseil financier personnalisé.")
    horizon_reco = st.slider("Horizon de projection (séances)", min_value=5, max_value=60, value=20, step=5, key="reco_horizon")
    reco_df = analyser_recommandations(df, horizon_reco)
    if reco_df.empty:
        st.info("Au moins 10 séances par titre sont nécessaires pour générer un signal.")
    else:
        ordre_reco = {'ACHAT À ÉTUDIER': 0, 'CONSERVER / SURVEILLER': 1, 'VENTE À ÉTUDIER': 2}
        reco_df['_ordre'] = reco_df['Signal'].map(ordre_reco)
        reco_df = reco_df.sort_values(['_ordre', 'Score'], ascending=[True, False]).drop(columns='_ordre')
        filtre_reco = st.multiselect("Afficher les signaux", list(ordre_reco), default=list(ordre_reco), key="reco_filtre")
        affichage_reco = reco_df[reco_df['Signal'].isin(filtre_reco)]
        a, b, c = st.columns(3)
        a.metric("Achats à étudier", int((reco_df['Signal'] == 'ACHAT À ÉTUDIER').sum()))
        b.metric("À surveiller", int((reco_df['Signal'] == 'CONSERVER / SURVEILLER').sum()))
        c.metric("Ventes à étudier", int((reco_df['Signal'] == 'VENTE À ÉTUDIER').sum()))
        st.dataframe(affichage_reco, hide_index=True, use_container_width=True, height=360)
        st.info("Méthode : score = 45 % rendement sur 20 séances + 45 % projection de tendance − 10 % volatilité. Les projections sont plafonnées à ±20 %.")


with tab_seance:
    st.markdown("### 📅 Données par séance de cotation")

    col_d, col_s, col_q = st.columns([2,2,2])
    with col_d:
        date_opts = [dt.strftime('%d/%m/%Y') for dt in all_dates]
        sel_date_str = st.selectbox("Date", date_opts, key="seance_date")
        sel_date = pd.to_datetime(sel_date_str, format='%d/%m/%Y')
    with col_s:
        sect_opts = ['Tous'] + sorted(df['secteur'].dropna().unique())
        sel_sect  = st.selectbox("Secteur", sect_opts, key="seance_sect")
    with col_q:
        search = st.text_input("Recherche symbole/titre", placeholder="SLBC, TOTALENERGIES…", key="seance_search")

    day_df = df[df['date_dt'] == sel_date].copy()
    if sel_sect != 'Tous': day_df = day_df[day_df['secteur']==sel_sect]
    if search:
        mask = (day_df['symbole'].str.contains(search.upper(),na=False) |
                day_df['titre'].str.contains(search,case=False,na=False))
        day_df = day_df[mask]

    disp_cols = {
        'symbole':'Symbole','titre':'Titre','secteur':'Secteur',
        'ouverture':'Ouv.','cloture':'Clôt.','variation_jour_pct':'Var. %',
        'volume':'Volume','valeur_fcfa':'Valeur (FCFA)',
        'variation_annee_pct':'Var. Année %','dividende_net':'Dividende',
        'rendement_net_pct':'Rdt. Net %','per':'PER'
    }
    disp = day_df[[c for c in disp_cols if c in day_df.columns]].rename(columns=disp_cols)
    st.caption(f"**{len(disp)}** titres pour le {sel_date_str}")

    def _style_var(v):
        if pd.isna(v): return 'color:#7A8699'
        return 'color:#2ECC71;font-weight:600' if v>0 else ('color:#E74C3C;font-weight:600' if v<0 else 'color:#7A8699')

    styled = disp.style
    for c in ['Var. %','Var. Année %','Rdt. Net %']:
        if c in disp.columns: styled = styled.applymap(_style_var, subset=[c])

    st.dataframe(styled, hide_index=True, use_container_width=True, height=520)


# ════════════════════════════════════════════════════════════════════════════
# ONGLET 3 — PAR TITRE
# ════════════════════════════════════════════════════════════════════════════
with tab_titre:
    st.markdown("### 🔍 Analyse détaillée par titre")

    col_sym, col_yr = st.columns([2,2])
    with col_sym:
        sel_sym = st.selectbox("Sélectionner un titre", avail_syms, key="titre_sym")
    with col_yr:
        yr_filter = st.selectbox("Filtrer par année", ['Toutes']+[str(y) for y in all_years], key="titre_yr")

    sym_df = df[df['symbole']==sel_sym].sort_values('date_dt')
    if yr_filter != 'Toutes':
        sym_df = sym_df[sym_df['date_dt'].dt.year==int(yr_filter)]

    if sym_df.empty:
        st.info("Aucune donnée pour ce titre / cette période.")
    else:
        last = sym_df.iloc[-1]
        first = sym_df.iloc[0]
        perf = (last['cloture']-first['cloture'])/first['cloture']*100 if first['cloture'] else None
        var_max = sym_df['cloture'].max()
        var_min = sym_df['cloture'].min()

        st.markdown(f"""
        <div class="brvm-card">
            <div style="display:flex;justify-content:space-between;align-items:flex-start;flex-wrap:wrap;gap:12px;">
                <div>
                    <div style="font-family:'Space Mono',monospace;color:#C9A84C;font-size:1.4rem;font-weight:700;">{sel_sym}</div>
                    <div style="color:#7A8699;font-size:.82rem;margin-top:2px;">{last.get('titre','')}</div>
                    <div style="margin-top:6px;"><span class="badge badge-info">{last.get('secteur','')}</span></div>
                </div>
                <div style="text-align:right;">
                    <div style="font-family:'Space Mono',monospace;color:#DDE3EF;font-size:1.8rem;line-height:1;">
                        {last['cloture']:,.0f}
                        <span style="color:#7A8699;font-size:.75rem;"> FCFA</span>
                    </div>
                    <div style="margin-top:4px;">
                        {badge(last.get('variation_jour_pct'))} {badge(perf)}
                    </div>
                </div>
            </div>
            <div style="display:grid;grid-template-columns:repeat(5,1fr);gap:8px;margin-top:14px;">
                <div style="background:var(--surface3);border-radius:7px;padding:9px;text-align:center;">
                    <div style="color:#7A8699;font-size:.62rem;text-transform:uppercase;">PER</div>
                    <div style="color:#DDE3EF;font-family:'Space Mono',monospace;font-size:.95rem;">{last.get('per') or '—'}</div>
                </div>
                <div style="background:var(--surface3);border-radius:7px;padding:9px;text-align:center;">
                    <div style="color:#7A8699;font-size:.62rem;text-transform:uppercase;">Rdt. Net</div>
                    <div style="color:#2ECC71;font-family:'Space Mono',monospace;font-size:.95rem;">{f"{last['rendement_net_pct']:.2f}%" if pd.notna(last.get('rendement_net_pct')) else '—'}</div>
                </div>
                <div style="background:var(--surface3);border-radius:7px;padding:9px;text-align:center;">
                    <div style="color:#7A8699;font-size:.62rem;text-transform:uppercase;">Plus haut</div>
                    <div style="color:#DDE3EF;font-family:'Space Mono',monospace;font-size:.95rem;">{var_max:,.0f}</div>
                </div>
                <div style="background:var(--surface3);border-radius:7px;padding:9px;text-align:center;">
                    <div style="color:#7A8699;font-size:.62rem;text-transform:uppercase;">Plus bas</div>
                    <div style="color:#DDE3EF;font-family:'Space Mono',monospace;font-size:.95rem;">{var_min:,.0f}</div>
                </div>
                <div style="background:var(--surface3);border-radius:7px;padding:9px;text-align:center;">
                    <div style="color:#7A8699;font-size:.62rem;text-transform:uppercase;">Séances</div>
                    <div style="color:#DDE3EF;font-family:'Space Mono',monospace;font-size:.95rem;">{len(sym_df)}</div>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        try:
            import plotly.graph_objects as go
            fig = go.Figure()
            # Candlestick si ouverture dispo, sinon ligne
            if sym_df['ouverture'].notna().sum() > len(sym_df)//2:
                fig.add_trace(go.Candlestick(
                    x=sym_df['date_dt'], open=sym_df['ouverture'],
                    high=sym_df[['ouverture','cloture']].max(axis=1),
                    low=sym_df[['ouverture','cloture']].min(axis=1),
                    close=sym_df['cloture'], name=sel_sym,
                    increasing=dict(fillcolor='rgba(46,204,113,.7)', line=dict(color='#2ECC71')),
                    decreasing=dict(fillcolor='rgba(231,76,60,.7)', line=dict(color='#E74C3C')),
                ))
            else:
                fig.add_trace(go.Scatter(
                    x=sym_df['date_dt'], y=sym_df['cloture'], mode='lines+markers',
                    line=dict(color='#C9A84C',width=2), marker=dict(size=4),
                    fill='tozeroy', fillcolor='rgba(201,168,76,.05)'
                ))
            fig.update_layout(
                paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                xaxis=dict(color='#7A8699', gridcolor='rgba(201,168,76,.07)', title='Date',
                           rangeslider=dict(visible=False)),
                yaxis=dict(color='#7A8699', gridcolor='rgba(201,168,76,.07)', title='FCFA'),
                margin=dict(l=0,r=0,t=20,b=40), height=300, font=dict(family='DM Sans'), showlegend=False
            )
            st.plotly_chart(fig, use_container_width=True)

            # Volume
            fig2 = go.Figure(go.Bar(
                x=sym_df['date_dt'], y=sym_df['volume'],
                marker=dict(color='rgba(59,130,246,.65)', line=dict(color='#3B82F6',width=.5))
            ))
            fig2.update_layout(
                paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                xaxis=dict(color='#7A8699', gridcolor='rgba(201,168,76,.05)'),
                yaxis=dict(color='#7A8699', gridcolor='rgba(201,168,76,.05)', title='Volume'),
                margin=dict(l=0,r=0,t=10,b=30), height=160
            )
            st.plotly_chart(fig2, use_container_width=True)
        except ImportError:
            st.line_chart(sym_df.set_index('date_dt')['cloture'])

        # Tableau historique
        st.markdown("#### 📋 Historique complet")
        hist = sym_df[['date','ouverture','cloture','variation_jour_pct','volume','valeur_fcfa','per']].copy()
        hist.columns = ['Date','Ouv.','Clôt.','Var. %','Volume','Valeur FCFA','PER']
        hist = hist.sort_values('Date', ascending=False)
        st.dataframe(hist, hide_index=True, use_container_width=True, height=320)


# ════════════════════════════════════════════════════════════════════════════
# ONGLET 4 — SÉRIES TEMPORELLES
# ════════════════════════════════════════════════════════════════════════════
with tab_ts:
    st.markdown("### 📈 Comparaison multi-titres")

    col_s1, col_s2, col_s3 = st.columns([3,1,1])
    with col_s1:
        defaults = [s for s in ['ETIT','SLBC','ORAC','BOAC'] if s in avail_syms][:3]
        sel_syms = st.multiselect("Titres à comparer (max 8)", avail_syms, default=defaults, max_selections=8)
    with col_s2:
        metric_map = {'cloture':'Cours Clôture','volume':'Volume',
                      'variation_jour_pct':'Var. Jour %','valeur_fcfa':'Valeur transigée'}
        metric = st.selectbox("Indicateur", list(metric_map.keys()), format_func=lambda x: metric_map[x])
    with col_s3:
        normalize = st.checkbox("Base 100", value=False)
        yr_range = st.multiselect("Années", [str(y) for y in all_years], default=[str(all_years[0])] if all_years else [])

    if sel_syms:
        filt_df = df.copy()
        if yr_range:
            filt_df = filt_df[filt_df['date_dt'].dt.year.astype(str).isin(yr_range)]

        try:
            import plotly.graph_objects as go
            PALETTE = ['#C9A84C','#3B82F6','#2ECC71','#E74C3C','#A855F7','#F39C12','#06B6D4','#EC4899']
            fig = go.Figure()
            for i, sym in enumerate(sel_syms):
                sd = filt_df[filt_df['symbole']==sym].sort_values('date_dt')
                if sd.empty: continue
                y = sd[metric].values.astype(float).copy()
                if normalize:
                    first_valid = y[~pd.isna(y)]
                    if len(first_valid) and first_valid[0] != 0:
                        y = y / first_valid[0] * 100
                fig.add_trace(go.Scatter(
                    x=sd['date_dt'], y=y, mode='lines', name=sym,
                    line=dict(color=PALETTE[i%len(PALETTE)], width=2),
                ))
            fig.update_layout(
                paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                xaxis=dict(color='#7A8699', gridcolor='rgba(201,168,76,.07)'),
                yaxis=dict(color='#7A8699', gridcolor='rgba(201,168,76,.07)',
                           title='Base 100' if normalize else metric_map[metric]),
                margin=dict(l=0,r=0,t=20,b=40), height=400,
                legend=dict(bgcolor='rgba(14,21,37,.85)', bordercolor='rgba(201,168,76,.3)',
                            borderwidth=1, font=dict(color='#DDE3EF',size=11)),
                hovermode='x unified'
            )
            st.plotly_chart(fig, use_container_width=True)
        except ImportError:
            piv = filt_df[filt_df['symbole'].isin(sel_syms)].pivot_table(
                index='date_dt', columns='symbole', values=metric)
            st.line_chart(piv)

        # Tableau de performance
        st.markdown("#### 📊 Tableau comparatif")
        rows_perf = []
        for sym in sel_syms:
            sd = filt_df[filt_df['symbole']==sym].sort_values('date_dt')
            if sd.empty: continue
            f, l = sd.iloc[0], sd.iloc[-1]
            perf = (l['cloture']-f['cloture'])/f['cloture']*100 if f['cloture'] else None
            rows_perf.append({
                'Symbole': sym, 'Titre': str(l.get('titre',''))[:28],
                'Dernier': f"{l['cloture']:,.0f} FCFA",
                'Var. Jour': f"{l.get('variation_jour_pct',0):+.2f}%" if pd.notna(l.get('variation_jour_pct')) else '—',
                'Perf. Période': f"{perf:+.2f}%" if perf else '—',
                'Vol. Moyen': fmt_num(sd['volume'].mean()),
                'PER': l.get('per','—'),
                'Rdt. Net': f"{l.get('rendement_net_pct',0):.2f}%" if pd.notna(l.get('rendement_net_pct')) else '—',
            })
        if rows_perf:
            st.dataframe(pd.DataFrame(rows_perf), hide_index=True, use_container_width=True)
    else:
        st.info("Sélectionnez au moins un titre.")


# ════════════════════════════════════════════════════════════════════════════
# ONGLET 5 — HISTORIQUE ANNUEL
# ════════════════════════════════════════════════════════════════════════════
with tab_histo:
    st.markdown("### 🗓️ Vue historique par année")

    col_hy, col_hm = st.columns([1,3])
    with col_hy:
        sel_yr = st.selectbox("Année", [str(y) for y in all_years], key="histo_yr")

    yr_df = df[df['date_dt'].dt.year==int(sel_yr)]

    if yr_df.empty:
        st.info(f"Aucune donnée pour {sel_yr}.")
    else:
        n_sess = yr_df['date_dt'].nunique()
        n_rec  = len(yr_df)
        avg_vol= yr_df['volume'].mean()
        tot_val= yr_df['valeur_fcfa'].sum()

        # KPIs année
        k1,k2,k3,k4 = st.columns(4)
        with k1: st.metric(f"Séances {sel_yr}", n_sess)
        with k2: st.metric("Enregistrements", f"{n_rec:,}")
        with k3: st.metric("Vol. moyen/titre", fmt_num(avg_vol))
        with k4: st.metric("Valeur totale échangée", fmt_num(tot_val," FCFA"))

        # Calendrier mensuel
        st.markdown("#### 📆 Séances disponibles par mois")
        monthly_yr = yr_df.groupby(yr_df['date_dt'].dt.month).agg(
            séances=('date_dt','nunique'),
            titres=('symbole','nunique'),
            vol_moyen=('volume','mean'),
            val_totale=('valeur_fcfa','sum')
        ).reset_index()
        monthly_yr['Mois'] = monthly_yr['date_dt'].apply(lambda m: datetime(int(sel_yr),m,1).strftime('%B'))
        monthly_yr = monthly_yr.rename(columns={'séances':'Séances','titres':'Titres',
                                                  'vol_moyen':'Vol. Moyen','val_totale':'Val. Totale FCFA'})
        monthly_yr['Vol. Moyen'] = monthly_yr['Vol. Moyen'].apply(fmt_num)
        monthly_yr['Val. Totale FCFA'] = monthly_yr['Val. Totale FCFA'].apply(fmt_num)
        st.dataframe(monthly_yr[['Mois','Séances','Titres','Vol. Moyen','Val. Totale FCFA']],
                     hide_index=True, use_container_width=True)

        # Top performers de l'année
        st.markdown("#### 🏆 Meilleures & pires performances de l'année")
        perf_rows = []
        for sym in yr_df['symbole'].unique():
            sd = yr_df[yr_df['symbole']==sym].sort_values('date_dt')
            if len(sd)<2: continue
            f, l = sd.iloc[0], sd.iloc[-1]
            if f['cloture'] and f['cloture']>0:
                perf = (l['cloture']-f['cloture'])/f['cloture']*100
                perf_rows.append({'Symbole':sym,'Titre':str(l.get('titre',''))[:25],
                                   'Cours déb.':f['cloture'],'Cours fin':l['cloture'],
                                   'Perf. Année %':round(perf,2),
                                   'Vol. Total':sd['volume'].sum()})
        if perf_rows:
            perf_df = pd.DataFrame(perf_rows).sort_values('Perf. Année %', ascending=False)
            col_top, col_bot = st.columns(2)
            with col_top:
                st.markdown("**Top 10 hausses**")
                st.dataframe(perf_df.head(10)[['Symbole','Titre','Cours déb.','Cours fin','Perf. Année %']],
                             hide_index=True, use_container_width=True, height=300)
            with col_bot:
                st.markdown("**Top 10 baisses**")
                st.dataframe(perf_df.tail(10)[['Symbole','Titre','Cours déb.','Cours fin','Perf. Année %']].sort_values('Perf. Année %'),
                             hide_index=True, use_container_width=True, height=300)

        # Graphique évolution indice synthétique (cours moyen pondéré)
        st.markdown("#### 📈 Évolution cours moyen pondéré — marché entier")
        daily_avg = yr_df.groupby('date_dt').agg(
            cours_moy=('cloture','mean'), vol_total=('volume','sum')
        ).reset_index()
        try:
            import plotly.graph_objects as go
            fig = go.Figure()
            fig.add_trace(go.Scatter(
                x=daily_avg['date_dt'], y=daily_avg['cours_moy'],
                mode='lines', line=dict(color='#C9A84C',width=2),
                fill='tozeroy', fillcolor='rgba(201,168,76,.04)'
            ))
            fig.update_layout(
                paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                xaxis=dict(color='#7A8699', gridcolor='rgba(201,168,76,.07)'),
                yaxis=dict(color='#7A8699', gridcolor='rgba(201,168,76,.07)', title='Cours moy. FCFA'),
                margin=dict(l=0,r=0,t=10,b=30), height=220
            )
            st.plotly_chart(fig, use_container_width=True)
        except ImportError:
            st.line_chart(daily_avg.set_index('date_dt')['cours_moy'])

        # Statut extraction de l'année
        log_dict = st.session_state.extract_log_dict
        all_cands = get_candidate_dates(int(sel_yr))
        n_ok_yr   = sum(1 for c in all_cands if log_dict.get(c,{}).get('status')=='ok')
        n_nf_yr   = sum(1 for c in all_cands if log_dict.get(c,{}).get('status')=='not_found')
        n_pend_yr = len(all_cands) - n_ok_yr - n_nf_yr
        pct_yr = n_ok_yr/len(all_cands)*100 if all_cands else 0

        st.markdown("#### 📊 Couverture extraction")
        st.markdown(f"""
        <div class="brvm-card" style="padding:14px;">
            <div style="display:flex;justify-content:space-between;margin-bottom:8px;">
                <span style="color:#7A8699;font-size:.75rem;">{sel_yr} — {n_ok_yr}/{len(all_cands)} dates</span>
                <span style="color:var(--gold);font-family:'Space Mono',monospace;font-size:.82rem;">{pct_yr:.1f}%</span>
            </div>
            <div style="background:var(--surface);border-radius:4px;height:8px;overflow:hidden;margin-bottom:8px;">
                <div style="background:linear-gradient(90deg,#C9A84C,#E8C96B);width:{pct_yr:.1f}%;height:100%;"></div>
            </div>
            <div style="display:flex;gap:16px;font-size:.72rem;color:#7A8699;">
                <span><span class="dot dot-g"></span>{n_ok_yr} extraits</span>
                <span><span class="dot dot-r"></span>{n_nf_yr} non trouvés (fériés/clôtures)</span>
                <span><span class="dot dot-o"></span>{n_pend_yr} en attente</span>
            </div>
        </div>
        """, unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════════════════════
# ONGLET 6 — EXPORT
# ════════════════════════════════════════════════════════════════════════════
with tab_export:
    st.markdown("### 💾 Export et téléchargement")

    # Filtres export
    st.markdown("#### 🎛️ Filtres")
    col_ef1, col_ef2, col_ef3 = st.columns(3)
    with col_ef1:
        exp_yrs = st.multiselect("Années", [str(y) for y in all_years], default=[str(all_years[0])], key='exp_yrs')
    with col_ef2:
        exp_syms = st.multiselect("Symboles (vide = tous)", avail_syms, key='exp_syms')
    with col_ef3:
        exp_sects = st.multiselect("Secteurs (vide = tous)", sorted(df['secteur'].dropna().unique()), key='exp_sects')

    exp_df = df.copy()
    if exp_yrs:  exp_df = exp_df[exp_df['date_dt'].dt.year.astype(str).isin(exp_yrs)]
    if exp_syms: exp_df = exp_df[exp_df['symbole'].isin(exp_syms)]
    if exp_sects:exp_df = exp_df[exp_df['secteur'].isin(exp_sects)]

    st.caption(f"**{len(exp_df):,}** lignes · **{exp_df['symbole'].nunique()}** titres · **{exp_df['date_dt'].nunique()}** séances")

    col_e1, col_e2, col_e3 = st.columns(3)

    # ── CSV brut ──────────────────────────────────────────────────────
    with col_e1:
        st.markdown("#### 📄 CSV Brut")
        st.caption("Toutes les colonnes, une ligne par (date, titre)")
        exp_out = exp_df.drop(columns=['date_dt'], errors='ignore').copy()
        csv_raw = exp_out.to_csv(index=False, sep=';', decimal=',', encoding='utf-8-sig')
        st.download_button(
            "⬇️ CSV Brut",
            data=csv_raw.encode('utf-8-sig'),
            file_name=f"BRVM_brut_{'_'.join(exp_yrs) if exp_yrs else 'all'}_{datetime.now().strftime('%Y%m%d')}.csv",
            mime="text/csv", use_container_width=True
        )

    # ── CSV Pivot cours clôture ───────────────────────────────────────
    with col_e2:
        st.markdown("#### 📊 CSV Pivot Clôture")
        st.caption("Format : Date, ABJC, BICB, … (une col. par titre)")
        piv = exp_df.pivot_table(index='date', columns='symbole', values='cloture', aggfunc='last').reset_index()
        ordered = ['date'] + [s for s in ALL_SYMBOLS if s in piv.columns]
        piv = piv[[c for c in ordered if c in piv.columns]]
        csv_piv = piv.to_csv(index=False, sep=',')
        st.download_button(
            "⬇️ CSV Pivot",
            data=csv_piv.encode('utf-8-sig'),
            file_name=f"BRVM_pivot_{'_'.join(exp_yrs) if exp_yrs else 'all'}_{datetime.now().strftime('%Y%m%d')}.csv",
            mime="text/csv", use_container_width=True
        )
        # Aperçu
        st.caption(f"Format : {piv.shape[0]} séances × {piv.shape[1]-1} titres")

    # ── Excel multi-feuilles ──────────────────────────────────────────
    with col_e3:
        st.markdown("#### 📗 Excel complet")
        st.caption("4 feuilles : données, pivot, stats, couverture")
        if st.button("🔧 Générer Excel", use_container_width=True, key="gen_excel"):
            with st.spinner("Génération..."):
                exp_out2 = exp_df.copy()
                xl_bytes = generate_excel_workbook(exp_out2)
                buf = io.BytesIO(xl_bytes)
                st.download_button(
                    "⬇️ Excel",
                    data=xl_bytes,
                    file_name=f"BRVM_{'_'.join(exp_yrs) if exp_yrs else 'all'}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    use_container_width=True
                )

    # Aperçu tableau
    st.markdown("---")
    st.markdown("#### 👁️ Aperçu")
    n_prev = st.slider("Lignes", 20, 500, 100, key='prev_slider')
    prev_df = exp_out.sort_values(['date','symbole'], ascending=[False,True]).head(n_prev)

    def _style_v(v):
        if pd.isna(v): return 'color:#7A8699'
        return 'color:#2ECC71;font-weight:600' if v>0 else ('color:#E74C3C;font-weight:600' if v<0 else 'color:#7A8699')

    styled_prev = prev_df.style
    for c in ['variation_jour_pct','variation_annee_pct']:
        if c in prev_df.columns: styled_prev = styled_prev.applymap(_style_v, subset=[c])
    st.dataframe(styled_prev, hide_index=True, use_container_width=True, height=400)


# ── Footer ──────────────────────────────────────────────────────────────────
st.markdown("""
<div style="margin-top:40px;padding:16px 0;border-top:1px solid rgba(201,168,76,.12);text-align:center;">
    <span style="color:#7A8699;font-size:.72rem;font-family:'Space Mono',monospace;">
        BRVM DATA EXTRACTOR v3 · Données BRVM · bfin.brvm.org · Usage académique &amp; professionnel
    </span>
</div>
""", unsafe_allow_html=True)


if c in disp.columns:
    if hasattr(styled, "map"):
        styled = styled.map(_style_var, subset=[c])
    else:
        styled = styled.applymap(_style_var, subset=[c])
